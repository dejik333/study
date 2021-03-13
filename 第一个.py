import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import time, os
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill, colors
import 全局变量


def main():
    def 桌面路径函数():
        return os.path.join(os.path.expanduser("~"), 'Desktop')
    桌面路径 = str(桌面路径函数())

    开始时间 = time.perf_counter()
    print('开始！！')
    print('正在加载文件,慢的话就是文件太大')

    原始表格 = openpyxl.load_workbook(r'%s\C类原始表格.xlsx' % 桌面路径)
    原始sheet = 原始表格.active

    最终表格 = openpyxl.Workbook()
    最终sheet = 最终表格.active
    最大行数 = 原始sheet.max_row

    print('文件加载完毕喽~~~~ 核对数据')
    加载文件时间 = time.perf_counter()

    系统表3 = openpyxl.load_workbook(r'%s\C类表3.xlsx' % 桌面路径)
    表3 = 系统表3.active

    def 得数值(值):
        try:
            值 = float(值)
            return 值
        except:
            return 0

    def 加边框(目标sheet):
        最终最大行数 = 最终sheet.max_row
        最终最大列数 = 最终sheet.max_column
        thin = Side(border_style="thin", color="000000")  # 边框样式，颜色
        border = Border(left=thin, right=thin, top=thin, bottom=thin)  # 边框的位置
        for 行 in range(1, 最终最大行数 + 1):
            for 列 in range(1, 最终最大列数 + 1):
                目标sheet.cell(row=行, column=列).border = border

    def 单位名称替换(单位老名称):
        单位新名称 = ''
        if 单位老名称 == '北京外企(江西)人力资源服务有限公司':
            单位新名称 = '北京外企江西公司'
        elif 单位老名称 == '北京外企德科人力资源服务苏州有限公司':
            单位新名称 = '北京外企苏州德科'
        elif 单位老名称 == '北京外企人力资源服务安徽有限公司':
            单位新名称 = '北京外企安徽公司'
        elif 单位老名称 == '北京外企人力资源服务福建有限公司':
            单位新名称 = '北京外企福建公司'
        elif 单位老名称 == '北京外企人力资源服务福建有限公司福州分公司':
            单位新名称 = '北京外企福建福州分公司'
        elif 单位老名称 == '北京外企人力资源服务河南有限公司':
            单位新名称 = '北京外企河南公司'
        elif 单位老名称 == '北京外企人力资源服务湖北有限公司':
            单位新名称 = '北京外企湖北公司'
        elif 单位老名称 == '北京外企人力资源服务济南有限公司':
            单位新名称 = '北京外企济南公司'
        elif 单位老名称 == '北京外企人力资源服务江苏有限公司':
            单位新名称 = '北京外企江苏公司'
        elif 单位老名称 == '北京外企人力资源服务江苏有限公司无锡分公司':
            单位新名称 = '北京外企江苏无锡分公司'
        elif 单位老名称 == '北京外企人力资源服务青岛有限公司':
            单位新名称 = '北京外企青岛公司'
        elif 单位老名称 == '北京外企人力资源服务有限公司广西分公司':
            单位新名称 = '北京外企广西分公司'
        elif 单位老名称 == '北京外企人力资源服务有限公司贵州分公司':
            单位新名称 = '北京外企贵州分公司'
        elif 单位老名称 == '北京外企人力资源服务有限公司宁波分公司':
            单位新名称 = '北京外企宁波分公司'
        elif 单位老名称 == '北京外企人力资源服务有限公司天津武清分公司':
            单位新名称 = '北京外企天津武清分公司'
        elif 单位老名称 == '北京外企人力资源服务有限公司通州分公司':
            单位新名称 = '北京外企通州公司'
        elif 单位老名称 == '北京外企人力资源服务云南有限公司':
            单位新名称 = '北京外企云南公司'
        elif 单位老名称 == '广东方胜人力资源服务有限公司':
            单位新名称 = '北京外企广东方胜'
        elif 单位老名称 == '河北方胜人力资源服务有限公司':
            单位新名称 = '北京外企河北方胜'
        elif 单位老名称 == '山西方胜人力资源服务有限公司':
            单位新名称 = '北京外企山西方胜'
        elif 单位老名称 == '四川方胜人力资源服务有限公司':
            单位新名称 = '北京外企四川方胜'
        elif 单位老名称 == '浙江外企德科人力资源服务有限公司':
            单位新名称 = '北京外企浙江德科'
        elif 单位老名称 == '重庆外企德科人力资源服务有限公司':
            单位新名称 = '北京外企重庆德科'
        elif 单位老名称 in '北京外企人力资源服务宁波有限公司':
            单位新名称 = '北京外企宁波公司'
        return 单位新名称

    def 加颜色(列数, 颜色):
        if 颜色 == '绿色':
            颜色 = "66CD00"
        if 颜色 == '深黄色':
            颜色 = "FFD39B"
        if 颜色 == '黄色':
            颜色 = "FFFF00"

        最大行数 = 最终sheet.max_row
        for 行数 in range(1, 最大行数 + 1):
            最终sheet.cell(row=行数, column=column_index_from_string(列数)).fill = PatternFill("solid", fgColor=颜色)

    def 求和(某列, 对应列, 需要的单位):
        总和 = 0.0
        for E in range(2, 最大行数 + 1):
            if 原始sheet.cell(row=E, column=column_index_from_string(某列)).value == 需要的单位:
                对应列的值 = 得数值(原始sheet.cell(row=E, column=column_index_from_string(对应列)).value)
                总和 += 对应列的值
        print('%s 在原表里的总额计算为 %s' % (需要的单位, str(总和)))
        return 总和

    不要的单位 = ['北京外企人力资源服务有限公司']
    global 结果反馈
    结果反馈 = ''

    def 单位汇报(老单位名称):
        单位新名称 = 单位名称替换(老单位名称)
        总额 = 得数值(求和('R', 'AS', 老单位名称))

        for E in range(1, 30):
            global 结果反馈
            if ((单位新名称 == 表3.cell(row=E, column=column_index_from_string('A')).value) or
                    (老单位名称 == 表3.cell(row=E, column=column_index_from_string('A')).value)):

                系统总额 = 表3.cell(row=E, column=column_index_from_string('B')).value
                if (总额 - 系统总额) > 1:
                    print(老单位名称 + '---------总额不符，计算总额是%f，系统总额是%f，'
                                 '差了%f' % (总额, 系统总额, (总额 - 系统总额)))

                    结果反馈 += '\n' + 单位新名称 + '\n'\
                           '计算总额是%d，系统总额是%d，' \
                                   '差了%d' % (总额, 系统总额, (总额 - 系统总额))
                    全局变量.set_value('结果反馈', 结果反馈)

                    不要的单位.append(老单位名称)
                else:
                    print('%s 总额无误,表3中的正确值为 %s' % (老单位名称, 系统总额))
                break

    单位汇报('北京外企(江西)人力资源服务有限公司')
    单位汇报('北京外企德科人力资源服务苏州有限公司')
    单位汇报('北京外企人力资源服务安徽有限公司')
    单位汇报('北京外企人力资源服务福建有限公司')
    单位汇报('北京外企人力资源服务福建有限公司福州分公司 ')
    单位汇报('北京外企人力资源服务河南有限公司')
    单位汇报('北京外企人力资源服务湖北有限公司')
    单位汇报('北京外企人力资源服务济南有限公司')
    单位汇报('北京外企人力资源服务江苏有限公司')
    单位汇报('北京外企人力资源服务江苏有限公司无锡分公司')
    单位汇报('北京外企人力资源服务青岛有限公司 ')
    单位汇报('北京外企人力资源服务有限公司广西分公司')
    单位汇报('北京外企人力资源服务有限公司贵州分公司')
    单位汇报('北京外企人力资源服务有限公司宁波分公司')
    单位汇报('北京外企人力资源服务有限公司天津武清分公司')
    单位汇报('北京外企人力资源服务有限公司通州分公司 ')
    单位汇报('北京外企人力资源服务云南有限公司')
    单位汇报('广东方胜人力资源服务有限公司')
    单位汇报('山西方胜人力资源服务有限公司')
    单位汇报('四川方胜人力资源服务有限公司')
    单位汇报('浙江外企德科人力资源服务有限公司')
    单位汇报('重庆外企德科人力资源服务有限公司')
    单位汇报('北京外企人力资源服务宁波有限公司')

    print('以上单位将不包含在新表中—————————————————————————————————————————————————————————————————————————————————————————————')
    print('开始数据运算------------------------------------------------------------------------------------------------------')
    核对完成时间 = time.perf_counter()

    # A
    最终sheet['A1'] = '账单年月'
    for E in range(2, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            当前时间 = time.strftime('%Y%m', time.localtime())
            最终sheet.cell(row=E, column=1).value = 当前时间

    # B
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            最终sheet.cell(row=E, column=2).value = 原始sheet.cell(row=E, column=column_index_from_string('C')).value

    # C
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            最终sheet.cell(row=E, column=3).value = 原始sheet.cell(row=E, column=column_index_from_string('D')).value

    # D
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            最终sheet.cell(row=E, column=4).value = 原始sheet.cell(row=E, column=column_index_from_string('F')).value

    # E
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            最终sheet.cell(row=E, column=5).value = 原始sheet.cell(row=E, column=column_index_from_string('H')).value

    # F
    最终sheet['F1'] = '委托单位'
    for E in range(2, 最大行数 + 1):
        单位老名称 = 原始sheet.cell(row=E, column=column_index_from_string('R')).value
        if 单位老名称 not in 不要的单位:
            单位新名称 = 单位名称替换(单位老名称)
            最终sheet.cell(row=E, column=6).value = 单位新名称

    # G
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            最终sheet.cell(row=E, column=7).value = 原始sheet.cell(row=E, column=column_index_from_string('M')).value

    # H
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            最终sheet.cell(row=E, column=8).value = 原始sheet.cell(row=E, column=column_index_from_string('U')).value

    # I
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            最终sheet.cell(row=E, column=9).value = 原始sheet.cell(row=E, column=column_index_from_string('AD')).value

    # J
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            最终sheet.cell(row=E, column=10).value = 原始sheet.cell(row=E, column=column_index_from_string('AE')).value

    # K
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            最终sheet.cell(row=E, column=11).value = 原始sheet.cell(row=E, column=column_index_from_string('AS')).value

    # K
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            最终sheet.cell(row=E, column=11).value = 原始sheet.cell(row=E, column=column_index_from_string('AS')).value

    # L
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            if 原始sheet.cell(row=E, column=column_index_from_string('BS')).value != '托收':
                最终sheet.cell(row=E, column=12).value = 原始sheet.cell(row=E, column=column_index_from_string('BJ')).value
            else:
                最终sheet.cell(row=E, column=12).value = None

    # N
    最终sheet['N1'] = '养老利息'
    # print('以下备注不包含养老,所以未处理:')
    # for E in range(2, 最大行数 + 1):
    #     if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
    #         临时收费备注 = 原始sheet.cell(row=E, column=column_index_from_string('AR')).value
    #         if 临时收费备注 is not None:
    #             if '养老' in 临时收费备注:
    #                 最终sheet.cell(row=E, column=14).value = 原始sheet.cell(row=E, column=column_index_from_string('AV')).value
    #             else:
    #                 print('         第%s行，备注内容为  %s' % (E, 临时收费备注.replace("\n", "")))

    # M
    最终sheet['M1'] = '养老保险缴纳额'
    for E in range(2, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            if 原始sheet.cell(row=E, column=column_index_from_string('BS')).value != '托收':
                if 原始sheet.cell(row=E, column=column_index_from_string('BS')).value is not None:
                    最终sheet.cell(row=E, column=13).value = 得数值(
                        原始sheet.cell(row=E, column=column_index_from_string('BR')).value) - (
                                                               得数值(原始sheet.cell(row=E, column=column_index_from_string(
                                                                   'AV')).value))
            else:
                最终sheet.cell(row=E, column=13).value = None

    # O
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            if 原始sheet.cell(row=E, column=column_index_from_string('CC')).value != '托收':
                最终sheet.cell(row=E, column=15).value = 原始sheet.cell(row=E, column=column_index_from_string('BT')).value
            else:
                最终sheet.cell(row=E, column=15).value = None

    # P
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            if 原始sheet.cell(row=E, column=column_index_from_string('CC')).value != '托收':
                最终sheet.cell(row=E, column=16).value = 原始sheet.cell(row=E, column=column_index_from_string('CB')).value
            else:
                最终sheet.cell(row=E, column=16).value = None

    # Q
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            if 原始sheet.cell(row=E, column=column_index_from_string('CM')).value != '托收':
                最终sheet.cell(row=E, column=17).value = 原始sheet.cell(row=E, column=column_index_from_string('CD')).value
            else:
                最终sheet.cell(row=E, column=17).value = None

    # R
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            if 原始sheet.cell(row=E, column=column_index_from_string('CM')).value != '托收':
                # if 原始sheet.cell(row=E, column=column_index_from_string('CM')).value is not None:
                最终sheet.cell(row=E, column=18).value = 原始sheet.cell(row=E, column=column_index_from_string('CL')).value
            else:
                最终sheet.cell(row=E, column=18).value = None

    # S
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            if 原始sheet.cell(row=E, column=column_index_from_string('CW')).value != '托收':
                最终sheet.cell(row=E, column=19).value = 原始sheet.cell(row=E, column=column_index_from_string('CN')).value
            else:
                最终sheet.cell(row=E, column=19).value = None

    # T
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            if 原始sheet.cell(row=E, column=column_index_from_string('CW')).value != '托收':
                最终sheet.cell(row=E, column=20).value = 原始sheet.cell(row=E, column=column_index_from_string('CV')).value
            else:
                最终sheet.cell(row=E, column=20).value = None

    # U
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            if 原始sheet.cell(row=E, column=column_index_from_string('DG')).value != '托收':
                最终sheet.cell(row=E, column=21).value = 原始sheet.cell(row=E, column=column_index_from_string('CX')).value
            else:
                最终sheet.cell(row=E, column=21).value = None

    # V
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            if 原始sheet.cell(row=E, column=column_index_from_string('DG')).value != '托收':
                最终sheet.cell(row=E, column=22).value = 原始sheet.cell(row=E, column=column_index_from_string('DF')).value
            else:
                最终sheet.cell(row=E, column=22).value = None

    # W
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            if 原始sheet.cell(row=E, column=column_index_from_string('DQ')).value != '托收':
                最终sheet.cell(row=E, column=23).value = 原始sheet.cell(row=E, column=column_index_from_string('DH')).value
            else:
                最终sheet.cell(row=E, column=23).value = None

    # X
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            if 原始sheet.cell(row=E, column=column_index_from_string('DQ')).value != '托收':
                最终sheet.cell(row=E, column=24).value = 原始sheet.cell(row=E, column=column_index_from_string('DP')).value
            else:
                最终sheet.cell(row=E, column=24).value = None

    # Y
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            if 原始sheet.cell(row=E, column=column_index_from_string('EA')).value != '托收':
                最终sheet.cell(row=E, column=25).value = 原始sheet.cell(row=E, column=column_index_from_string('DR')).value
            else:
                最终sheet.cell(row=E, column=25).value = None

    # Z
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            if 原始sheet.cell(row=E, column=column_index_from_string('EA')).value != '托收':
                最终sheet.cell(row=E, column=26).value = 原始sheet.cell(row=E, column=column_index_from_string('DZ')).value
            else:
                最终sheet.cell(row=E, column=26).value = None

    # AA
    最终sheet['AA1'] = '合计'

    for 行数 in range(2, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            最终sheet.cell(row=行数, column=column_index_from_string('AA')).value = (
                    '=SUM(I%s,M%s,N%s,P%s,R%s,T%s,V%s,X%s,Z%s)' %
                    (str(行数), str(行数), str(行数), str(行数), str(行数), str(行数), str(行数), str(行数), str(行数)))

    # AB
    最终sheet['AB1'] = '核对'

    for 行数 in range(2, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            最终sheet.cell(row=行数, column=column_index_from_string('AB')).value = (
                    '=SUM(AA%s-K%s)' %
                    (str(行数), str(行数)))

    # AC
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            临时收费备注 = 原始sheet.cell(row=E, column=column_index_from_string('AR')).value
            if 临时收费备注 is not None:
                最终sheet.cell(row=E, column=29).value = 原始sheet.cell(row=E, column=column_index_from_string('AR')).value
            # else:
            #     最终sheet.cell(row=E, column=29).value = None

    # AD
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            临时费用总额 = 原始sheet.cell(row=E, column=column_index_from_string('AF')).value
            if 临时费用总额 is not None:
                最终sheet.cell(row=E, column=30).value = 原始sheet.cell(row=E, column=column_index_from_string('AF')).value

            # else:
            #     最终sheet.cell(row=E, column=30).value = None

    # AE
    for E in range(1, 最大行数 + 1):
        if 原始sheet.cell(row=E, column=column_index_from_string('R')).value not in 不要的单位:
            内容 = 原始sheet.cell(row=E, column=column_index_from_string('AQ')).value
            if 内容 is not None:
                最终sheet.cell(row=E, column=31).value = 原始sheet.cell(row=E, column=column_index_from_string('AQ')).value

    计算时间 = time.perf_counter()

    # print('数据转移完毕，现在删除表格中的空行------------------------------------------------------------------------------------')

    # 删除空行后的表格 = openpyxl.Workbook()
    # 删除空行后的表格sheet = 删除空行后的表格.active
    # 新行 = 0
    # for 老行 in range(1, 最大行数+1):
    #     if 最终sheet.cell(row=老行, column=1).value is not None:
    #         新行 += 1
    #         for 列数 in range(1, 200):
    #             删除空行后的表格sheet.cell(row=新行, column=列数).value = 最终sheet.cell(row=老行, column=列数).value
    #         print('第%s行生成完毕' % 新行)
    # print('-----------------上色-------------')
    # 加颜色('I', '绿色')
    # 加颜色('J', '绿色')
    # 加颜色('K', '绿色')
    # 加颜色('N', '深黄色')
    # 加颜色('M', '黄色')
    # 加颜色('P', '黄色')
    # 加颜色('R', '黄色')
    # 加颜色('T', '黄色')
    # 加颜色('V', '黄色')
    # 加颜色('X', '黄色')
    # 加颜色('Z', '黄色')
    # 加边框(最终sheet)
    print('正在保存。。。。')
    最终表格.save(r'%s\C类结果.xlsx' % 桌面路径)

    print('==================完成=================== ')
    print('总用时为: %f  秒\n其中:' % (计算时间 - 开始时间))
    print('    加载文件用时: %f 秒' % (加载文件时间 - 开始时间))
    print('    核对总额用时: %f 秒' % (核对完成时间 - 加载文件时间))
    print('    计算   用时: %f 秒' % (计算时间 - 核对完成时间))
    # print('    删除北京用时: %f 秒' % (删除空行时间-计算时间))
    print('==================结束===================    请仔细检查，本程序不对结果负责，若领导追责，解释权归豆豆妈所有……')


if __name__ == "__main__":
    main()
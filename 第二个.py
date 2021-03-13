
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import time, os
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill, colors
import 全局变量


def main():
    def 桌面路径函数():
        return os.path.join(os.path.expanduser("~"), 'Desktop')
    桌面路径 = str(桌面路径函数())
    
    开始时间 = time.perf_counter()
    print('开始！！')
    print('正在加载文件,慢的话就是文件太大')

    # 把表格里的公式都计算为结果后打开
    表1文件 = openpyxl.load_workbook(r'%s\上海德科表1.xlsx' % 桌面路径, data_only=True)
    表2文件 = openpyxl.load_workbook(r'%s\上海德科表2.xlsx' % 桌面路径, data_only=True)

    表1 = 表1文件['账单明细']
    表2 = 表2文件['账单明细']
    表2客户sheet = 表2文件['客户费用']

    最终表格 = openpyxl.Workbook()
    最终sheet = 最终表格.active

    表1最大行数 = 表1.max_row
    表2最大行数 = 表2.max_row
    表2客户sheet最大行数 = 表2客户sheet.max_row
    # print(表1最大行数, 表2最大行数, 表2客户sheet最大行数)
    最终最大行数 = 表1最大行数 + 表2最大行数 + 表2客户sheet最大行数 - 5

    def 得数值(值):
        try:
            值 = float(值)
            return 值
        except:
            return 0

    def 加边框(目标sheet):
        最终最大行数 = 最终sheet.max_row
        最终最大列数 = 最终sheet.max_column
        thin = Side(border_style="thin", color="000000")  # 边框样式，颜色
        border = Border(left=thin, right=thin, top=thin, bottom=thin)  # 边框的位置
        for 行 in range(1, 最终最大行数+1):
            for 列 in range(1, 最终最大列数+1):
                目标sheet.cell(row=行, column=列).border = border

    def 加颜色(列数, 颜色):
        if 颜色 == '绿色':
            颜色 = "66CD00"
        if 颜色 == '深黄色':
            颜色 = "FFD39B"
        if 颜色 == '黄色':
            颜色 = "FFFF00"
        最大行数 = 最终最大行数

        for 行数 in range(1, 最大行数+1):
            最终sheet.cell(row=行数, column=column_index_from_string(列数)).fill = PatternFill("solid", fgColor=颜色)

    def 下面粘贴(最终表格列, 表1列=None, 表2列=None, 表3列=None):
        最终表格列 = column_index_from_string(最终表格列)
        if 表1列 is not None:
            for 表1行数 in range(3, 表1最大行数+1):
                # 从表1第3行开始
                # if 表1列 is None:
                #     最终sheet.cell(row=表1行数-1, column=最终表格列).value = None
                # else:
                最终sheet.cell(row=表1行数-1, column=最终表格列).value = 表1.cell(row=表1行数, column=column_index_from_string(表1列)).value
        if 表2列 is not None:
            for 表2行数 in range(3, 表2最大行数+1):
                # 从表2的第3行开始
                # if 表2列 is None:
                #     最终sheet.cell(row=表1最大行数 + 表2行数-3, column=最终表格列).value = ''
                # else:
                最终sheet.cell(row=表1最大行数 + 表2行数-3, column=最终表格列).value = 表2.cell(row=表2行数, column=column_index_from_string(表2列)).value
        if 表3列 is not None:
            for 表3行数 in range(2, 表2客户sheet最大行数+1):
                # 从表3的第2行开始
                # if 表3列 is None:
                #     最终sheet.cell(row=表1最大行数 + 表2最大行数 + 表3行数-6, column=最终表格列).value = ''
                # else:
                最终sheet.cell(row=表1最大行数 + 表2最大行数+表3行数-6, column=最终表格列).value = 表2客户sheet.cell(
                    row=表3行数, column=column_index_from_string(表3列)).value

    打开文件时间 = time.perf_counter()
    print('文件加载完毕喽~~~~ ')

    最终sheet['A1'] = '账单年月'
    for E in range(2, 最终最大行数+1):
        当前时间 = time.strftime('%Y%m', time.localtime())
        最终sheet.cell(row=E, column=1).value = 当前时间

    最终sheet['B1'] = '雇员唯一号'
    下面粘贴('B', 'D', 'D')

    最终sheet['C1'] = '雇员姓名'
    下面粘贴('C', 'C', 'C')

    最终sheet['D1'] = '身份证号'
    下面粘贴('D', 'E', 'E')

    最终sheet['E1'] = '投保地'
    下面粘贴('E', 'P', 'P')

    最终sheet['F1'] = '委托单位'
    for E in range(2, 最终最大行数):
        最终sheet.cell(row=E, column=6).value = '北京外企上海德科'

    最终sheet['G1'] = '业务客户'
    下面粘贴('G', 'N', 'N', 'D')

    最终sheet['H1'] = '业务年月'
    下面粘贴('H', 'T', 'T')

    最终sheet['I1'] = '服务费缴纳合计'
    # 下面粘贴('I', 表1列=None, 表2列='BN', 表3列='H')
    for E in range(2, 表2最大行数):
        if 表2.cell(row=E + 1, column=column_index_from_string('BN')).value is None:
            表2.cell(row=E + 1, column=column_index_from_string('BN')).value = '0'

        if 表2.cell(row=E + 1, column=column_index_from_string('BO')).value is None:
            表2.cell(row=E + 1, column=column_index_from_string('BO')).value = '0'
        # 临时用AI记录一下求和的值
        表2.cell(row=E + 1, column=column_index_from_string('AI')).value = (
                float(表2.cell(row=E + 1, column=column_index_from_string('BN')).value) +
                float(表2.cell(row=E + 1, column=column_index_from_string('BO')).value))

    下面粘贴('I', 表1列=None, 表2列='AI', 表3列='H')

    最终sheet['J1'] = '福利产品总额'

    最终sheet['K1'] = '实付总金额'
    # 下面粘贴('K', 表1列='BR', 表2列='BS', 表3列='H')
    # 把表1 BR列公式的值计算出来
    for E in range(2, 表1最大行数):
        表1.cell(row=E + 1, column=column_index_from_string('BP')).value = (
                得数值(表1.cell(row=E + 1, column=column_index_from_string('X')).value) +
                得数值(表1.cell(row=E + 1, column=column_index_from_string('AA')).value) +
                得数值(表1.cell(row=E + 1, column=column_index_from_string('AE')).value) +
                得数值(表1.cell(row=E + 1, column=column_index_from_string('AH')).value) +
                得数值(表1.cell(row=E + 1, column=column_index_from_string('AL')).value) +
                得数值(表1.cell(row=E + 1, column=column_index_from_string('AO')).value) +
                得数值(表1.cell(row=E + 1, column=column_index_from_string('AS')).value) +
                得数值(表1.cell(row=E + 1, column=column_index_from_string('AW')).value) +
                得数值(表1.cell(row=E + 1, column=column_index_from_string('BA')).value) +
                得数值(表1.cell(row=E + 1, column=column_index_from_string('BD')).value) +
                得数值(表1.cell(row=E + 1, column=column_index_from_string('BI')).value) +
                得数值(表1.cell(row=E + 1, column=column_index_from_string('BL')).value)
        )

        if 表1.cell(row=E + 1, column=column_index_from_string('BP')).value is None:
            表1.cell(row=E + 1, column=column_index_from_string('BP')).value = '0'

        if 表1.cell(row=E + 1, column=column_index_from_string('BQ')).value is None:
            表1.cell(row=E + 1, column=column_index_from_string('BQ')).value = '0'

        表1.cell(row=E+1, column=column_index_from_string('BR')).value = (
                float(表1.cell(row=E + 1, column=column_index_from_string('BP')).value) +
                float(表1.cell(row=E + 1, column=column_index_from_string('BQ')).value))

    下面粘贴('K', 表1列='BR', 表3列='H')
    # 把I列数据复制过来
    for E in range(表1最大行数, 表1最大行数+表2最大行数):
        最终sheet.cell(row=E, column=column_index_from_string('K')).value = (
                最终sheet.cell(row=E, column=column_index_from_string('I')).value)

    最终sheet['L1'] = '养老保险企业基数'
    下面粘贴('L', 'V')

    最终sheet['M1'] = '养老保险缴纳额'
    for E in range(2, 表1最大行数):

        if 表1.cell(row=E + 1, column=column_index_from_string('X')).value is None:
            表1.cell(row=E + 1, column=column_index_from_string('X')).value = '0'

        if 表1.cell(row=E + 1, column=column_index_from_string('AA')).value is None:
            表1.cell(row=E + 1, column=column_index_from_string('AA')).value = '0'

        最终sheet.cell(row=E, column=column_index_from_string('M')).value = (
                float(表1.cell(row=E+1, column=column_index_from_string('X')).value) +
                float(表1.cell(row=E+1, column=column_index_from_string('AA')).value))

    最终sheet['N1'] = '养老利息'

    最终sheet['O1'] = '失业保险企业基数'
    下面粘贴('O', 'AJ')

    最终sheet['P1'] = '失业保险缴纳额'
    for E in range(2, 表1最大行数):
        if 表1.cell(row=E + 1, column=column_index_from_string('AL')).value is None:
            表1.cell(row=E + 1, column=column_index_from_string('AL')).value = '0'

        if 表1.cell(row=E + 1, column=column_index_from_string('AO')).value is None:
            表1.cell(row=E + 1, column=column_index_from_string('AO')).value = '0'

        最终sheet.cell(row=E, column=column_index_from_string('P')).value = (
                float(表1.cell(row=E+1, column=column_index_from_string('AL')).value) +
                float(表1.cell(row=E+1, column=column_index_from_string('AO')).value))

    最终sheet['Q1'] = '工伤保险企业基数'
    下面粘贴('Q', 'AQ')

    最终sheet['R1'] = '工伤保险缴纳额'
    下面粘贴('R', 'AS')

    最终sheet['S1'] = '生育保险企业基数'
    下面粘贴('S', 'AU')

    最终sheet['T1'] = '生育保险缴纳额'
    下面粘贴('T', 'AW')

    最终sheet['U1'] = '基本医疗保险企业基数'
    下面粘贴('U', 'AC')

    最终sheet['V1'] = '基本医疗保险缴纳额'
    for E in range(2, 表1最大行数):
        if 表1.cell(row=E + 1, column=column_index_from_string('AE')).value is None:
            表1.cell(row=E + 1, column=column_index_from_string('AE')).value = '0'

        if 表1.cell(row=E + 1, column=column_index_from_string('AH')).value is None:
            表1.cell(row=E + 1, column=column_index_from_string('AH')).value = '0'

        最终sheet.cell(row=E, column=column_index_from_string('V')).value = (
                float(表1.cell(row=E+1, column=column_index_from_string('AE')).value) +
                float(表1.cell(row=E+1, column=column_index_from_string('AH')).value))

    最终sheet['W1'] = '大病附加保险企业基数'
    下面粘贴('W', 'AY')

    最终sheet['X1'] = '大病附加保险缴纳额'
    for E in range(2, 表1最大行数):
        if 表1.cell(row=E + 1, column=column_index_from_string('BA')).value is None:
            表1.cell(row=E + 1, column=column_index_from_string('BA')).value = '0'

        if 表1.cell(row=E + 1, column=column_index_from_string('BD')).value is None:
            表1.cell(row=E + 1, column=column_index_from_string('BD')).value = '0'

        最终sheet.cell(row=E, column=column_index_from_string('X')).value = (
                float(表1.cell(row=E+1, column=column_index_from_string('BA')).value) +
                float(表1.cell(row=E+1, column=column_index_from_string('BD')).value))

    最终sheet['Y1'] = '住房公积金企业基数'
    下面粘贴('Y', 'BG')

    最终sheet['Z1'] = '住房公积金缴纳额'

    for E in range(2, 表1最大行数):
        if 表1.cell(row=E+1, column=column_index_from_string('BI')).value is None:
            表1.cell(row=E+1, column=column_index_from_string('BI')).value = '0'

        if 表1.cell(row=E+1, column=column_index_from_string('BL')).value is None:
            表1.cell(row=E+1, column=column_index_from_string('BL')).value = '0'

        表1.cell(row=E+1, column=column_index_from_string('BM')).value = (
                float(表1.cell(row=E + 1, column=column_index_from_string('BI')).value) +
                float(表1.cell(row=E + 1, column=column_index_from_string('BL')).value))

    下面粘贴('Z', 'BM')

    最终sheet['AA1'] = '合计'

    for 行数 in range(2, 表1最大行数+表2最大行数 + 表2客户sheet最大行数):
            最终sheet.cell(row=行数, column=column_index_from_string('AA')).value = (
                    '=SUM(I%s,M%s,N%s,P%s,R%s,T%s,V%s,X%s,Z%s)' %
                    (str(行数), str(行数), str(行数), str(行数), str(行数), str(行数), str(行数), str(行数), str(行数)))
    # for E in range(2, 最终最大行数+1):
    #
    #     I = 最终sheet.cell(row=E, column=column_index_from_string('I')).value
    #     M = 最终sheet.cell(row=E, column=column_index_from_string('M')).value
    #     N = 最终sheet.cell(row=E, column=column_index_from_string('N')).value
    #     P = 最终sheet.cell(row=E, column=column_index_from_string('P')).value
    #     R = 最终sheet.cell(row=E, column=column_index_from_string('R')).value
    #     T = 最终sheet.cell(row=E, column=column_index_from_string('T')).value
    #     V = 最终sheet.cell(row=E, column=column_index_from_string('V')).value
    #     X = 最终sheet.cell(row=E, column=column_index_from_string('X')).value
    #     Z = 最终sheet.cell(row=E, column=column_index_from_string('Z')).value
    #     if I is None:
    #          I = '0'
    #     if M is None:
    #         M = '0'
    #     if N is None:
    #         N = '0'
    #     if P is None:
    #         P = '0'
    #     if R is None:
    #         R = '0'
    #     if T is None:
    #         T = '0'
    #     if V is None:
    #         V = '0'
    #     if X is None:
    #         X = '0'
    #     if Z is None:
    #         Z = '0'
    #     最终sheet.cell(row=E, column=27).value = (
    #                                             得数值(I) + 得数值(M) + 得数值(N) + 得数值(P) + 得数值(R) +
    #                                             得数值(T) + 得数值(V) + 得数值(X) + 得数值(Z)
    #                                             )

    最终sheet['AB1'] = '核对'

    for 行数 in range(2, 表1最大行数 + 表2最大行数 + 10):
        最终sheet.cell(row=行数, column=column_index_from_string('AB')).value = (
                '=SUM(AA%s-K%s)' %
                (str(行数), str(行数)))

    最终sheet['AC1'] = '备注'
    下面粘贴('AC', 'BS')

    计算时间 = time.perf_counter()
    # print('计算完成，加颜色')
    # 加颜色('I', '绿色')
    # 加颜色('J', '绿色')
    # 加颜色('K', '绿色')
    # 加颜色('N', '深黄色')
    # 加颜色('M', '黄色')
    # 加颜色('P', '黄色')
    # 加颜色('R', '黄色')
    # 加颜色('T', '黄色')
    # 加颜色('V', '黄色')
    # 加颜色('X', '黄色')
    # 加颜色('Z', '黄色')
    # 加边框(最终sheet)
    最终表格.save(r'%s\德科结果.xlsx' % 桌面路径)
    # 全局变量.set_value('结果反馈', '')

    print('==================完成=================== ')
    print('总用时为: %f  秒\n其中:' % (计算时间-开始时间))
    print('    加载文件用时: %f 秒' % (打开文件时间-开始时间))
    print('    核对总额用时: %f 秒' % (计算时间-打开文件时间))
    print('==================结束===================    请仔细检查，本程序不对结果负责，若领导追责，解释权归豆豆妈所有……')


if __name__ =="__main__":
    main()
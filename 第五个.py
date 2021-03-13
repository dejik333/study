
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import time, os
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill, colors
import 全局变量


def main():
    def 桌面路径函数():
        return os.path.join(os.path.expanduser("~"), 'Desktop')
    桌面路径 = str(桌面路径函数())

    开始时间 = time.perf_counter()

    原始文件 = openpyxl.load_workbook(r'%s\苏州原始表格.xlsx' % 桌面路径, data_only=True)
    原始表 = 原始文件['个人费用']
    最大行数 = 原始表.max_row

    最终表格 = openpyxl.Workbook()
    最终sheet = 最终表格.active

    def 加边框(目标sheet):
        最终最大行数 = 最终sheet.max_row
        最终最大列数 = 最终sheet.max_column
        thin = Side(border_style="thin", color="000000")  # 边框样式，颜色
        border = Border(left=thin, right=thin, top=thin, bottom=thin)  # 边框的位置
        for 行 in range(1, 最终最大行数+1):
            for 列 in range(1, 最终最大列数+1):
                目标sheet.cell(row=行, column=列).border = border

    def 加颜色(列数, 颜色):
        if 颜色 == '绿色':
            颜色 = "66CD00"
        if 颜色 == '深黄色':
            颜色 = "FFD39B"
        if 颜色 == '黄色':
            颜色 = "FFFF00"

        for 行数 in range(1, 最大行数+1):
            最终sheet.cell(row=行数, column=column_index_from_string(列数)).fill = PatternFill("solid", fgColor=颜色)

    def 转移(新列, 旧列):
        for 行数 in range(1, 最大行数+1):
            最终sheet.cell(row=行数, column=column_index_from_string(新列)).value = 原始表.cell(row=行数, column=column_index_from_string(旧列)).value

    def 得数值(值):
        try:
            值 = float(值)
            return 值
        except:
            return 0

    最终sheet['A1'] = '账单年月'
    for E in range(2, 最大行数+1):
        当前时间 = time.strftime('%Y%m', time.localtime())
        最终sheet.cell(row=E, column=1).value = 当前时间

    转移('B', 'C')
    转移('C', 'A')
    转移('D', 'B')
    转移('E', 'I')

    最终sheet['F1'] = '委托单位'
    for 行数 in range(2, 最大行数 + 1):
        最终sheet.cell(row=行数, column=column_index_from_string('F')).value = '北京外企苏州德科'

    转移('G', 'F')
    转移('H', 'N')

    最终sheet['I1'] = '服务费'
    for E in range(2, 最大行数+1):
        I = 原始表.cell(row=E, column=column_index_from_string('CR')).value
        M = 原始表.cell(row=E, column=column_index_from_string('CS')).value
        N = 原始表.cell(row=E, column=column_index_from_string('CQ')).value

        最终sheet.cell(row=E, column=column_index_from_string('I')).value = 得数值(I) + 得数值(M) + 得数值(N)

    最终sheet['J1'] = '福利产品'
    转移('K', 'DD')
    转移('L', 'Q')
    转移('M', 'Y')
    最终sheet['N1'] = '养老利息'
    转移('O', 'AM')
    转移('P', 'AU')
    转移('Q', 'AX')
    转移('R', 'BF')
    转移('S', 'BI')
    转移('T', 'BQ')
    转移('U', 'AB')
    转移('V', 'AJ')
    转移('W', 'BT')
    转移('X', 'CB')
    转移('Y', 'CG')
    转移('Z', 'CO')

    最终sheet['AA1'] = '合计'
    for 行数 in range(2, 最大行数+1):
        最终sheet.cell(row=行数, column=column_index_from_string('AA')).value = (
                '=SUM(I%s,M%s,N%s,P%s,R%s,T%s,V%s,X%s,Z%s)' %
                (str(行数), str(行数), str(行数), str(行数), str(行数), str(行数), str(行数), str(行数), str(行数)))

    加边框(最终sheet)

    最终sheet['AB1'] = '核对'

    for 行数 in range(2, 最大行数+1):
        最终sheet.cell(row=行数, column=column_index_from_string('AB')).value = (
                '=SUM(AA%s-K%s)' %
                (str(行数), str(行数)))

    转移('AD', 'DG')

    加颜色('I', '绿色')
    加颜色('J', '绿色')
    加颜色('K', '绿色')
    加颜色('N', '深黄色')
    加颜色('M', '黄色')
    加颜色('P', '黄色')
    加颜色('R', '黄色')
    加颜色('T', '黄色')
    加颜色('V', '黄色')
    加颜色('X', '黄色')
    加颜色('Z', '黄色')

    # 单立户代理清零
    for 行数 in range(2, 最大行数 + 1):
        # 养老
        if 原始表.cell(row=行数, column=column_index_from_string('Q')).value == '单立户代理':
            最终sheet.cell(row=行数, column=column_index_from_string('L')).value = ''
            最终sheet.cell(row=行数, column=column_index_from_string('M')).value = ''
        # 失业
        if 原始表.cell(row=行数, column=column_index_from_string('AM')).value == '单立户代理':
            最终sheet.cell(row=行数, column=column_index_from_string('O')).value = ''
            最终sheet.cell(row=行数, column=column_index_from_string('P')).value = ''
        # 工伤
        if 原始表.cell(row=行数, column=column_index_from_string('AX')).value == '单立户代理':
            最终sheet.cell(row=行数, column=column_index_from_string('Q')).value = ''
            最终sheet.cell(row=行数, column=column_index_from_string('R')).value = ''
        # 生育
        if 原始表.cell(row=行数, column=column_index_from_string('BI')).value == '单立户代理':
            最终sheet.cell(row=行数, column=column_index_from_string('S')).value = ''
            最终sheet.cell(row=行数, column=column_index_from_string('T')).value = ''
        # 基本医疗
        if 原始表.cell(row=行数, column=column_index_from_string('AB')).value == '单立户代理':
            最终sheet.cell(row=行数, column=column_index_from_string('U')).value = ''
            最终sheet.cell(row=行数, column=column_index_from_string('V')).value = ''
        # 大病
        if 原始表.cell(row=行数, column=column_index_from_string('BT')).value == '单立户代理':
            最终sheet.cell(row=行数, column=column_index_from_string('W')).value = ''
            最终sheet.cell(row=行数, column=column_index_from_string('X')).value = ''
        # 公积金
        if 原始表.cell(row=行数, column=column_index_from_string('CG')).value == '单立户代理':
            最终sheet.cell(row=行数, column=column_index_from_string('Y')).value = ''
            最终sheet.cell(row=行数, column=column_index_from_string('Z')).value = ''

    最终表格.save(r'%s\苏州结果.xlsx' % 桌面路径)
    完成时间 = time.perf_counter()
    # 全局变量.set_value('结果反馈', '')

    print('==============================这就完了======================================\n程序共用时: %f 秒' % (完成时间-开始时间))
    print('================================结束======================================== 请仔细检查，本程序不对结果负责，若领导追责，解释权归豆豆妈所有……')


if __name__ == "__main__":
    main()
